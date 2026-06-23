import sys
import io
import os
from datetime import datetime

if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

URL = "https://hydro.bg/bg/t1.php?ime=&gr=data/&gn=tablRekiB2017"
STATION_ID = "51880"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_FILE = os.path.join(SCRIPT_DIR, "STRYMON_MARINO_POLE_SCRAPER.xlsx")


def parse_bg_number(text):
    """Convert Bulgarian formatted number (comma as decimal) to float.
    Returns None for non-numeric values like 'n.a.'."""
    if text == "n.a." or not text.strip():
        return None
    return float(text.replace(".", "").replace(",", "."))


def fetch_hydro_data():
    resp = requests.get(URL, timeout=30)
    resp.encoding = "utf-8"
    soup = BeautifulSoup(resp.text, "html.parser")

    # Extract measurement date from the page header
    measurement_date = None
    h2 = soup.find("h2")
    if h2:
        text = h2.get_text()
        # Format: "12.06.2026 08:00 часа местно време"
        for part in text.split():
            if "." in part and len(part) == 10:
                try:
                    measurement_date = datetime.strptime(part, "%d.%m.%Y").date()
                    break
                except ValueError:
                    pass

    target_table = None
    for h3 in soup.find_all("h3"):
        if "Западнобеломорски" in h3.get_text():
            target_table = h3.find_next("table")
            break

    if not target_table:
        print("Could not find the Западнобеломорски басейн table.")
        return None

    for row in target_table.find_all("tr"):
        cells = row.find_all("td")
        if cells and cells[0].get_text(strip=True) == STATION_ID:
            river = cells[1].get_text(strip=True)
            station_name = cells[2].get_text(strip=True)
            q_min = cells[3].get_text(strip=True)
            q_avg = cells[4].get_text(strip=True)
            q_max = cells[5].get_text(strip=True)
            h = cells[6].get_text(strip=True)
            q = cells[7].get_text(strip=True)
            delta_h = cells[8].get_text(strip=True)

            print(f"Station: {STATION_ID} - {river}, {station_name}")
            print(f"Q [m3/s]:  {q}")
            print(f"Delta H [cm]: {delta_h}")
            print()
            print("Full row data:")
            print(f"  Q_min:  {q_min} m3/s")
            print(f"  Q_avg:  {q_avg} m3/s")
            print(f"  Q_max:  {q_max} m3/s")
            print(f"  H:      {h} cm")
            print(f"  Q:      {q} m3/s")
            print(f"  DeltaH: {delta_h} cm")
            return {
                "date": measurement_date,
                "q_min": q_min,
                "q_avg": q_avg,
                "q_max": q_max,
                "h": h,
                "q": q,
                "delta_h": delta_h,
            }

    print(f"Station {STATION_ID} not found.")
    return None


def create_workbook():
    """Create a new xlsx file with formatted headers."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Data"

    headers = ["Date", "Q [m3/s]", "H [cm]", "dH [cm]", "Q_min [m3/s]", "Q_avg [m3/s]", "Q_max [m3/s]"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Set column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16

    return wb


def save_to_excel(data):
    """Append data to the xlsx file. Skip if today's data already exists."""
    if data is None or data["date"] is None:
        print("No valid data to save.")
        return

    if os.path.exists(XLSX_FILE):
        wb = load_workbook(XLSX_FILE)
        ws = wb.active
    else:
        wb = create_workbook()
        ws = wb.active

    # Check if date already exists
    existing_dates = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            if isinstance(row[0], datetime):
                existing_dates.add(row[0].date())
            else:
                existing_dates.add(row[0])

    if data["date"] in existing_dates:
        print(f"Data for {data['date']} already exists. Skipping.")
        return

    # Append new row
    next_row = ws.max_row + 1
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    values = [
        data["date"],
        parse_bg_number(data["q"]),
        parse_bg_number(data["h"]),
        parse_bg_number(data["delta_h"]),
        parse_bg_number(data["q_min"]),
        parse_bg_number(data["q_avg"]),
        parse_bg_number(data["q_max"]),
    ]

    for col, val in enumerate(values, 1):
        cell = ws.cell(row=next_row, column=col, value=val)
        cell.border = thin_border
        if col == 1:
            cell.number_format = "YYYY-MM-DD"
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.number_format = "0.000"
            cell.alignment = Alignment(horizontal="right")

    wb.save(XLSX_FILE)
    print(f"Saved to {XLSX_FILE} (row {next_row - 1})")


if __name__ == "__main__":
    data = fetch_hydro_data()
    if data:
        save_to_excel(data)
